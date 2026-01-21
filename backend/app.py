"""
Backend API para Sistema de Pase de Lista con Huellas Digitales
Sistema para registro de asistencia de alumnos de escuela media superior
Usa WebAuthn API para leer huellas desde dispositivos móviles
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from webauthn_handler import WebAuthnHandler, WebAuthnDatabase
import base64
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile

app = Flask(__name__)
CORS(app)  # Permitir CORS para React

# Configuración
DB_FILE = "alumnos.json"
GRUPOS_FILE = "grupos.json"

# Instancia global del manejador WebAuthn
webauthn_handler = None
webauthn_db = None


def get_webauthn_handler():
    """Obtiene o crea la instancia del manejador WebAuthn"""
    global webauthn_handler
    if webauthn_handler is None:
        webauthn_handler = WebAuthnHandler(db_file=DB_FILE)
    return webauthn_handler


def get_webauthn_db():
    """Obtiene o crea la instancia de la base de datos"""
    global webauthn_db
    if webauthn_db is None:
        webauthn_db = WebAuthnDatabase(db_file=DB_FILE)
    return webauthn_db


def load_grupos():
    """Carga los grupos desde el archivo"""
    if os.path.exists(GRUPOS_FILE):
        try:
            with open(GRUPOS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_grupos(grupos):
    """Guarda los grupos en el archivo"""
    try:
        with open(GRUPOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(grupos, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error al guardar grupos: {e}")
        return False


@app.route('/api/health', methods=['GET'])
def health_check():
    """Verifica el estado del sistema"""
    db = get_webauthn_db()
    grupos = load_grupos()
    
    total_alumnos = len(db.fingerprints)
    
    return jsonify({
        'status': 'ok',
        'webauthn_disponible': True,
        'total_alumnos': total_alumnos,
        'total_grupos': len(grupos)
    })


@app.route('/api/dispositivo/estado', methods=['GET'])
def dispositivo_estado():
    """Obtiene el estado del sistema WebAuthn"""
    # WebAuthn está disponible si el navegador lo soporta (se verifica en el frontend)
    return jsonify({
        'webauthn_disponible': True,
        'dispositivo_disponible': True,  # Siempre disponible si el navegador soporta WebAuthn
        'mensaje': 'Usa el lector de huellas de tu dispositivo móvil'
    })


# ==================== GRUPOS ====================

@app.route('/api/grupos', methods=['GET'])
def listar_grupos():
    """Lista todos los grupos registrados"""
    grupos = load_grupos()
    
    # Agregar información adicional de cada grupo
    grupos_list = []
    for grupo_id, grupo_data in grupos.items():
        # Contar alumnos en este grupo
        db = get_webauthn_db()
        alumnos_en_grupo = sum(1 for a in db.fingerprints.values() 
                              if a.get('grupo_id') == grupo_id)
        
        grupos_list.append({
            'grupo_id': grupo_id,
            'nombre': grupo_data.get('nombre', ''),
            'carrera_tecnica': grupo_data.get('carrera_tecnica', ''),
            'created_at': grupo_data.get('created_at', ''),
            'total_alumnos': alumnos_en_grupo
        })
    
    return jsonify({
        'total': len(grupos_list),
        'grupos': grupos_list
    })


@app.route('/api/grupos', methods=['POST'])
def crear_grupo():
    """Crea un nuevo grupo"""
    data = request.json
    
    if not data or 'nombre' not in data:
        return jsonify({'error': 'Se requiere el nombre del grupo'}), 400
    
    nombre = data.get('nombre', '').strip()
    carrera_tecnica = data.get('carrera_tecnica', '').strip()
    
    if not nombre:
        return jsonify({'error': 'El nombre del grupo es requerido'}), 400
    
    grupos = load_grupos()
    
    # Generar ID único para el grupo
    grupo_id = f"GRP-{len(grupos) + 1:03d}"
    
    # Verificar que no exista un grupo con el mismo nombre
    for gid, gdata in grupos.items():
        if gdata.get('nombre', '').upper() == nombre.upper():
            return jsonify({'error': 'Ya existe un grupo con ese nombre'}), 400
    
    grupos[grupo_id] = {
        'nombre': nombre,
        'carrera_tecnica': carrera_tecnica,
        'created_at': datetime.now().isoformat()
    }
    
    if save_grupos(grupos):
        return jsonify({
            'message': 'Grupo creado exitosamente',
            'grupo_id': grupo_id,
            'nombre': nombre,
            'carrera_tecnica': carrera_tecnica
        }), 201
    else:
        return jsonify({'error': 'Error al guardar el grupo'}), 500


@app.route('/api/grupos/<grupo_id>', methods=['GET'])
def obtener_grupo(grupo_id):
    """Obtiene información de un grupo específico"""
    grupos = load_grupos()
    
    if grupo_id not in grupos:
        return jsonify({'error': 'Grupo no encontrado'}), 404
    
    grupo_data = grupos[grupo_id]
    db = get_webauthn_db()
    
    # Obtener alumnos del grupo
    alumnos = [
        {
            'user_id': uid,
            'name': data.get('name', ''),
            'tiene_huella': len(data.get('credentials', [])) > 0,
            'registered_at': data.get('registered_at', '')
        }
        for uid, data in db.fingerprints.items()
        if data.get('grupo_id') == grupo_id
    ]
    
    # Ordenar alfabéticamente por nombre
    alumnos.sort(key=lambda x: x['name'].upper())
    
    return jsonify({
        'grupo_id': grupo_id,
        'nombre': grupo_data.get('nombre', ''),
        'carrera_tecnica': grupo_data.get('carrera_tecnica', ''),
        'created_at': grupo_data.get('created_at', ''),
        'total_alumnos': len(alumnos),
        'alumnos': alumnos
    })


@app.route('/api/grupos/<grupo_id>', methods=['DELETE'])
def eliminar_grupo(grupo_id):
    """Elimina un grupo"""
    grupos = load_grupos()
    
    if grupo_id not in grupos:
        return jsonify({'error': 'Grupo no encontrado'}), 404
    
    # Verificar si tiene alumnos
    db = get_webauthn_db()
    alumnos_en_grupo = sum(1 for a in db.fingerprints.values() 
                          if a.get('grupo_id') == grupo_id)
    
    if alumnos_en_grupo > 0:
        return jsonify({
            'error': f'No se puede eliminar el grupo. Tiene {alumnos_en_grupo} alumno(s) registrado(s)'
        }), 400
    
    del grupos[grupo_id]
    
    if save_grupos(grupos):
        return jsonify({'message': 'Grupo eliminado exitosamente'})
    else:
        return jsonify({'error': 'Error al eliminar el grupo'}), 500


# ==================== ALUMNOS ====================

@app.route('/api/grupos/<grupo_id>/alumnos', methods=['GET'])
def listar_alumnos_grupo(grupo_id):
    """Lista todos los alumnos de un grupo específico"""
    grupos = load_grupos()
    
    if grupo_id not in grupos:
        return jsonify({'error': 'Grupo no encontrado'}), 404
    
    db = get_webauthn_db()
    alumnos = []
    
    for user_id, user_data in db.fingerprints.items():
        if user_data.get('grupo_id') == grupo_id:
            alumnos.append({
                'user_id': user_id,
                'name': user_data.get('name', 'Sin nombre'),
                'registered_at': user_data.get('registered_at', ''),
                'tiene_huella': len(user_data.get('credentials', [])) > 0,
                'huella_registrada_at': user_data.get('credentials', [{}])[-1].get('registered_at', '') if user_data.get('credentials') else ''
            })
    
    # Ordenar alfabéticamente por nombre
    alumnos.sort(key=lambda x: x['name'].upper())
    
    return jsonify({
        'grupo_id': grupo_id,
        'grupo_nombre': grupos[grupo_id].get('nombre', ''),
        'total': len(alumnos),
        'alumnos': alumnos
    })


@app.route('/api/grupos/<grupo_id>/alumnos', methods=['POST'])
def registrar_alumno_grupo(grupo_id):
    """Registra un nuevo alumno en un grupo"""
    grupos = load_grupos()
    
    if grupo_id not in grupos:
        return jsonify({'error': 'Grupo no encontrado'}), 404
    
    data = request.json
    
    if not data or 'name' not in data:
        return jsonify({'error': 'Se requiere el nombre del alumno'}), 400
    
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'error': 'El nombre del alumno es requerido'}), 400
    
    db = get_webauthn_db()
    
    # Generar ID único para el alumno
    alumnos_en_grupo = sum(1 for a in db.fingerprints.values() 
                          if a.get('grupo_id') == grupo_id)
    user_id = f"{grupo_id}-{alumnos_en_grupo + 1:03d}"
    
    # Registrar alumno sin credenciales aún
    if user_id not in db.fingerprints:
        db.fingerprints[user_id] = {
            'name': name,
            'registered_at': datetime.now().isoformat(),
            'credentials': [],
            'grupo_id': grupo_id
        }
    else:
        db.fingerprints[user_id]['name'] = name
        db.fingerprints[user_id]['grupo_id'] = grupo_id
    
    db.save_database()
    
    return jsonify({
        'message': 'Alumno registrado exitosamente',
        'user_id': user_id,
        'name': name,
        'grupo_id': grupo_id
    }), 201


@app.route('/api/alumnos/<alumno_id>/huella/challenge', methods=['POST'])
def crear_challenge_registro(alumno_id):
    """Crea un challenge para registrar la huella del alumno usando WebAuthn"""
    db = get_webauthn_db()
    
    # Verificar que el alumno existe
    if alumno_id not in db.fingerprints:
        return jsonify({'error': 'Alumno no encontrado. Regístralo primero.'}), 404
    
    handler = get_webauthn_handler()
    challenge = handler.create_challenge()
    
    return jsonify({
        'challenge': challenge,
        'user_id': alumno_id,
        'user_name': db.fingerprints[alumno_id].get('name', '')
    })


@app.route('/api/alumnos/<alumno_id>/huella', methods=['POST'])
def registrar_huella_alumno(alumno_id):
    """Registra la credencial WebAuthn del alumno"""
    db = get_webauthn_db()
    
    # Verificar que el alumno existe
    if alumno_id not in db.fingerprints:
        return jsonify({'error': 'Alumno no encontrado. Regístralo primero.'}), 404
    
    data = request.json
    
    if not data or 'credential_id' not in data or 'public_key' not in data:
        return jsonify({'error': 'Datos de credencial WebAuthn requeridos'}), 400
    
    credential_id = data.get('credential_id')
    public_key = data.get('public_key')
    
    # Registrar la credencial
    alumno_data = db.fingerprints[alumno_id]
    db.add_credential(alumno_id, credential_id, public_key, alumno_data.get('name', ''))
    
    return jsonify({
        'message': 'Huella registrada exitosamente',
        'user_id': alumno_id,
        'name': alumno_data.get('name', ''),
        'credential_id': credential_id
    })


@app.route('/api/alumnos/<alumno_id>', methods=['DELETE'])
def eliminar_alumno(alumno_id):
    """Elimina un alumno del sistema"""
    db = get_webauthn_db()
    
    if db.delete_user(alumno_id):
        return jsonify({'message': 'Alumno eliminado exitosamente'})
    else:
        return jsonify({'error': 'Alumno no encontrado'}), 404


@app.route('/api/alumnos/<alumno_id>/asistencias', methods=['GET'])
def obtener_asistencias(alumno_id):
    """Obtiene el historial de asistencias de un alumno"""
    db = get_webauthn_db()
    
    if alumno_id not in db.fingerprints:
        return jsonify({'error': 'Alumno no encontrado'}), 404
    
    alumno_data = db.fingerprints[alumno_id]
    asistencias = alumno_data.get('asistencias', [])
    
    return jsonify({
        'user_id': alumno_id,
        'name': alumno_data.get('name', ''),
        'total_asistencias': len(asistencias),
        'asistencias': asistencias
    })


@app.route('/api/alumnos/<alumno_id>/asistencias', methods=['POST'])
def registrar_asistencia_manual(alumno_id):
    """Registra asistencia manualmente (sin huella)"""
    db = get_webauthn_db()
    
    if alumno_id not in db.fingerprints:
        return jsonify({'error': 'Alumno no encontrado'}), 404
    
    data = request.json or {}
    alumno_data = db.fingerprints[alumno_id]
    
    # Crear registro de asistencia
    asistencia = {
        'user_id': alumno_id,
        'name': alumno_data.get('name', ''),
        'timestamp': datetime.now().isoformat(),
        'fecha': datetime.now().strftime('%Y-%m-%d'),
        'hora': datetime.now().strftime('%H:%M:%S'),
        'tipo': data.get('tipo', 'manual')
    }
    
    # Guardar asistencia
    if 'asistencias' not in alumno_data:
        alumno_data['asistencias'] = []
    alumno_data['asistencias'].append(asistencia)
    db.save_database()
    
    return jsonify({
        'message': 'Asistencia registrada exitosamente',
        'asistencia': asistencia
    }), 201


# ==================== ASISTENCIA ====================

@app.route('/api/asistencia/verificar/challenge', methods=['POST'])
def crear_challenge_verificacion():
    """Crea un challenge para verificar asistencia usando WebAuthn"""
    data = request.json or {}
    grupo_id = data.get('grupo_id')  # Opcional: filtrar por grupo
    
    handler = get_webauthn_handler()
    challenge = handler.create_challenge()
    
    # Obtener lista de credential IDs permitidos si se especifica grupo
    allowed_credentials = []
    if grupo_id:
        db = get_webauthn_db()
        for user_id, user_data in db.fingerprints.items():
            if user_data.get('grupo_id') == grupo_id:
                credentials = user_data.get('credentials', [])
                for cred in credentials:
                    allowed_credentials.append({
                        'id': cred['credential_id'],
                        'type': 'public-key'
                    })
    
    return jsonify({
        'challenge': challenge,
        'allowed_credentials': allowed_credentials if allowed_credentials else None
    })


@app.route('/api/asistencia/verificar', methods=['POST'])
def verificar_asistencia():
    """Verifica la asistencia usando credencial WebAuthn"""
    data = request.json or {}
    grupo_id = data.get('grupo_id')  # Opcional: filtrar por grupo
    credential_id = data.get('credential_id')
    
    if not credential_id:
        return jsonify({'error': 'Credencial ID requerido'}), 400
    
    db = get_webauthn_db()
    
    # Buscar usuario por credential_id
    user_id = db.find_user_by_credential_id(credential_id)
    
    if not user_id:
        return jsonify({
            'encontrado': False,
            'message': 'Huella no reconocida. Alumno no registrado.'
        })
    
    alumno_data = db.fingerprints[user_id]
    
    # Si se especificó un grupo, verificar que el alumno pertenezca a ese grupo
    if grupo_id and alumno_data.get('grupo_id') != grupo_id:
        return jsonify({
            'encontrado': False,
            'message': 'El alumno no pertenece a este grupo'
        })
    
    # Registrar asistencia
    asistencia = {
        'user_id': user_id,
        'name': alumno_data.get('name', ''),
        'timestamp': datetime.now().isoformat(),
        'fecha': datetime.now().strftime('%Y-%m-%d'),
        'hora': datetime.now().strftime('%H:%M:%S')
    }
    
    # Guardar asistencia
    if 'asistencias' not in alumno_data:
        alumno_data['asistencias'] = []
    alumno_data['asistencias'].append(asistencia)
    db.save_database()
    
    return jsonify({
        'encontrado': True,
        'alumno': {
            'user_id': user_id,
            'name': alumno_data.get('name', ''),
            'grupo_id': alumno_data.get('grupo_id', '')
        },
        'asistencia': asistencia
    })


# ==================== ESTADÍSTICAS ====================

@app.route('/api/estadisticas', methods=['GET'])
def obtener_estadisticas():
    """Obtiene estadísticas del sistema"""
    db = get_webauthn_db()
    grupos = load_grupos()
    
    total_alumnos = len(db.fingerprints)
    alumnos_con_huella = sum(1 for a in db.fingerprints.values() 
                            if len(a.get('credentials', [])) > 0)
    total_asistencias = sum(len(a.get('asistencias', [])) for a in db.fingerprints.values())
    
    # Estadísticas por grupo
    estadisticas_grupos = {}
    for grupo_id, grupo_data in grupos.items():
        alumnos_grupo = [a for a in db.fingerprints.values() if a.get('grupo_id') == grupo_id]
        estadisticas_grupos[grupo_id] = {
            'nombre': grupo_data.get('nombre', ''),
            'carrera_tecnica': grupo_data.get('carrera_tecnica', ''),
            'total_alumnos': len(alumnos_grupo),
            'alumnos_con_huella': sum(1 for a in alumnos_grupo 
                                     if len(a.get('credentials', [])) > 0),
            'total_asistencias': sum(len(a.get('asistencias', [])) for a in alumnos_grupo)
        }
    
    # Asistencias por fecha
    asistencias_por_fecha = {}
    for alumno_data in db.fingerprints.values():
        for asistencia in alumno_data.get('asistencias', []):
            fecha = asistencia.get('fecha', '')
            if fecha:
                asistencias_por_fecha[fecha] = asistencias_por_fecha.get(fecha, 0) + 1
    
    return jsonify({
        'total_alumnos': total_alumnos,
        'alumnos_con_huella': alumnos_con_huella,
        'alumnos_sin_huella': total_alumnos - alumnos_con_huella,
        'total_asistencias': total_asistencias,
        'total_grupos': len(grupos),
        'estadisticas_grupos': estadisticas_grupos,
        'asistencias_por_fecha': asistencias_por_fecha
    })


@app.route('/api/estadisticas/descargar-excel', methods=['GET'])
def descargar_excel_asistencias():
    """Genera y descarga un archivo Excel con las asistencias"""
    grupo_id = request.args.get('grupo_id', None)
    db = get_webauthn_db()
    grupos = load_grupos()
    
    # Obtener alumnos (del grupo específico o todos)
    if grupo_id:
        alumnos_data = [
            (user_id, data) for user_id, data in db.fingerprints.items()
            if data.get('grupo_id') == grupo_id
        ]
        grupo_info = grupos.get(grupo_id, {})
        nombre_archivo = f"asistencias_{grupo_info.get('nombre', grupo_id).replace(' ', '_')}.xlsx"
        titulo = f"Asistencias - {grupo_info.get('nombre', grupo_id)}"
    else:
        alumnos_data = list(db.fingerprints.items())
        nombre_archivo = "asistencias_todos_los_grupos.xlsx"
        titulo = "Asistencias - Todos los Grupos"
    
    # Ordenar alumnos alfabéticamente por nombre
    alumnos_data.sort(key=lambda x: x[1].get('name', '').upper())
    
    # Obtener todas las fechas únicas de todas las asistencias
    fechas_set = set()
    for user_id, alumno_data in alumnos_data:
        for asistencia in alumno_data.get('asistencias', []):
            fecha = asistencia.get('fecha', '')
            if fecha:
                fechas_set.add(fecha)
    
    # Ordenar fechas cronológicamente
    fechas_ordenadas = sorted(list(fechas_set))
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar título en la parte superior
    ws.insert_rows(1)
    ws.merge_cells(f'A1:{get_column_letter(len(fechas_ordenadas) + 1)}1')
    title_cell = ws['A1']
    title_cell.value = titulo
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Encabezado: Nombre del alumno (ahora en fila 2)
    ws['A2'] = 'Alumno'
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    ws['A2'].border = border_style
    
    # Encabezados: Fechas (ahora en fila 2)
    for idx, fecha in enumerate(fechas_ordenadas, start=2):
        col = get_column_letter(idx)
        cell = ws[f'{col}2']
        cell.value = fecha
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_style
    
    # Llenar datos de alumnos (empezando en fila 3)
    for row_idx, (user_id, alumno_data) in enumerate(alumnos_data, start=3):
        nombre = alumno_data.get('name', user_id)
        
        # Nombre del alumno
        cell = ws[f'A{row_idx}']
        cell.value = nombre
        cell.border = border_style
        cell.alignment = Alignment(vertical='center')
        
        # Obtener fechas de asistencia de este alumno
        fechas_asistencia = set()
        for asistencia in alumno_data.get('asistencias', []):
            fecha = asistencia.get('fecha', '')
            if fecha:
                fechas_asistencia.add(fecha)
        
        # Marcar asistencias por fecha
        for col_idx, fecha in enumerate(fechas_ordenadas, start=2):
            col = get_column_letter(col_idx)
            cell = ws[f'{col}{row_idx}']
            
            if fecha in fechas_asistencia:
                cell.value = '*'  # Asistió
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.value = '/'  # No asistió
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            cell.border = border_style
            cell.alignment = center_alignment
            cell.font = Font(size=12)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    for idx in range(2, len(fechas_ordenadas) + 2):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = 12
    
    # Guardar en un archivo temporal
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    print("=" * 60)
    print("  SISTEMA DE PASE DE LISTA - BACKEND API")
    print("=" * 60)
    print("  Usando WebAuthn API para lectura de huellas")
    print(f"Database: {DB_FILE}")
    print(f"Grupos: {GRUPOS_FILE}")
    print("\nIniciando servidor en http://localhost:5000")
    print("Endpoints disponibles:")
    print("  Grupos:")
    print("    GET    /api/grupos")
    print("    POST   /api/grupos")
    print("    GET    /api/grupos/<id>")
    print("    DELETE /api/grupos/<id>")
    print("  Alumnos:")
    print("    GET    /api/grupos/<id>/alumnos")
    print("    POST   /api/grupos/<id>/alumnos")
    print("    POST   /api/alumnos/<id>/huella/challenge")
    print("    POST   /api/alumnos/<id>/huella")
    print("    DELETE /api/alumnos/<id>")
    print("  Asistencia:")
    print("    POST   /api/asistencia/verificar/challenge")
    print("    POST   /api/asistencia/verificar")
    print("  Estadísticas:")
    print("    GET    /api/estadisticas")
    print("    GET    /api/estadisticas/descargar-excel?grupo_id=<id>")
    print("=" * 60)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
